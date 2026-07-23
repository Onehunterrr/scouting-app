import json, sys
from transfer_pathways import PATHWAYS
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule

FONT_NAME = "Arial"
BLUE = Font(name=FONT_NAME, size=10, color="0000FF")
BLACK = Font(name=FONT_NAME, size=10, color="000000")
GREEN = Font(name=FONT_NAME, size=10, color="008000")
LINK = Font(name=FONT_NAME, size=10, color="0563C1", underline="single")
HEADER_FONT = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F3A5F")
YELLOW_FILL = PatternFill("solid", fgColor="FFFF00")
TITLE_FONT = Font(name=FONT_NAME, size=16, bold=True, color="1F3A5F")
SUB_FONT = Font(name=FONT_NAME, size=10, italic=True, color="555555")
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

DATA_FILE = sys.argv[1] if len(sys.argv) > 1 else "players_current.json"
OUT_FILE = sys.argv[2] if len(sys.argv) > 2 else "Scouting_Model.xlsx"

players = json.load(open(DATA_FILE))
N = len(players)
last_row = 1 + N  # header row 1

wb = Workbook()

# ---------------------------------------------------------------- Read Me
ws = wb.active
ws.title = "Read Me"
ws.column_dimensions["A"].width = 100
ws["A1"] = "Global Lower-Tier Scouting Model"
ws["A1"].font = TITLE_FONT
ws["A2"] = "A statistical model for spotting players who are performing above their market visibility, worldwide."
ws["A2"].font = SUB_FONT

lines = [
    ("", ""),
    ("What this workbook does", "h"),
    (f"Player Data holds {N} SAMPLE players (fictional, ages 17-26, across {len(set(p['country'] for p in players))} countries on six continents) illustrating the schema. Scoring Model computes per-90 rates, ranks each player against peers at the same position, and produces an Undervalued Score: the gap between how well a player performs and how visible they already are in the market.", "p"),
    ("", ""),
    ("Color legend", "h"),
    ("Blue text = a value you type in or edit.", "p"),
    ("Black text = a formula that only uses cells on its own sheet.", "p"),
    ("Green text = a formula that pulls from another sheet.", "p"),
    ("Yellow fill = an assumption you can tune (weights, tier multipliers).", "p"),
    ("", ""),
    ("How to use it", "h"),
    ("1. Add players to Player Data. Position must be GK, DF, MF, or FW; Tier must be 2, 3, or 4.", "p"),
    ("2. Scoring Model updates automatically and ranks every player against others in the same position, worldwide.", "p"),
    ("3. Sort or filter Scoring Model by the Undervalued Score or Flag column (AutoFilter is enabled) to build a shortlist.", "p"),
    ("4. Tune the weights on Assumptions if you want to emphasize different skills.", "p"),
    ("", ""),
    ("Representation & contact", "h"),
    ("Player Data tracks Has Agent (Yes/No/Unknown), Contract Expires, a Club Contact Email (illustrative format), and a Preferred Contact Route. Players under 18 are always routed to the club's youth/academy office, never direct contact.", "p"),
    ("", ""),
    ("Verifying identity and age", "h"),
    ("Age Verification Source names the real national federation registry to check. Verify Age / Bio is a live Transfermarkt search link (not a fixed profile URL), so it never misattributes this sample data to a real person. Age fraud is a known risk in scouting outside the top divisions -- always confirm DOB against an official source.", "p"),
    ("", ""),
    ("Database freshness", "h"),
    ("Date Added / Last Updated track when a player entered the database and when their stats were last refreshed. This workbook is regenerated on a weekly cadence: existing players get a light stat refresh (simulating a week of match activity) and 10 new players are added. See weekly_update.py.", "p"),
    ("", ""),
    ("Tactical system fit", "h"),
    ("System Fit is a rules-based read of which tactical system a player's statistical profile suits best (High Press, Possession / Build-from-the-Back, Counter-Attack / Transition, Low Block / Park the Bus, Direct / Target Approach, or Flexible / Multi-System), derived from their percentiles and pass-vs-carry tendency. It's a heuristic based on sample data, not a scout's tactical assessment.", "p"),
    ("", ""),
    ("Realistic transfer targets", "h"),
    ("Transfer Pathway Countries lists the two realistic destination countries for each player's country, based on well-known lower-tier/global scouting corridors (e.g. Balkans -> Austria/Turkey, West Africa -> Belgium/Portugal). Move Type classifies whether a player's performance percentile supports a Step-Up move, a Lateral Move, or is best served staying and developing. The prototype app shows the full Top 5 Realistic Transfer Targets breakdown with illustrative (fictional) destination clubs -- these are model output, not real clubs' actual interest.", "p"),
    ("", ""),
    ("Goalkeeper scoring (revised)", "h"),
    ("Goalkeepers previously reused the outfield GA/Prog/Def formulas, which barely differentiated one keeper from another. Grounded in real goalkeeper-analytics research (PSxG+/-, save %, distribution accuracy, and FBref's #OPA sweeper-actions-outside-the-box metric), GK scoring now uses three keeper-specific factors instead: Shot-Stopping (saves minus a 1.5x goals-conceded penalty -- a simplified proxy for PSxG+/- since this dataset has no shot-level data), Distribution (pass completion %), and Sweeper Actions per 90. These slot into the same weighted-percentile structure as outfield players, just with goalkeeper-appropriate inputs.", "p"),
    ("", ""),
    ("How this differs from Wyscout / InStat", "h"),
    ("Wyscout (which absorbed InStat after Hudl's 2021 acquisition) is a strong, broad tool -- 250+ competitions, video plus data -- but it's built for club recruitment departments and priced/structured accordingly, and it's a general-purpose search-and-filter tool rather than one built around a specific discovery workflow. This model's differentiation is narrower and more specific: an explicit Undervalued Score (performance percentile minus market-visibility percentile, not just raw stats), representation-status tracking built into the core workflow (Has Agent, minor-safety contact routing) rather than bolted on, and a weekly-growing sample database designed for someone building a client base from scratch rather than a club with an existing scouting department and budget.", "p"),
    ("", ""),
    ("Important caveat", "h"),
    ("This model is a filter, not a verdict. It can't see injuries, attitude, or off-pitch risk. Treat a high score as \"go watch this player,\" never as \"sign this player.\"", "p"),
]
r = 3
for text, kind in lines:
    c = ws.cell(row=r, column=1, value=text)
    if kind == "h":
        c.font = Font(name=FONT_NAME, size=12, bold=True, color="1F3A5F")
    elif kind == "p":
        c.font = Font(name=FONT_NAME, size=10)
        ws.row_dimensions[r].height = 28
        c.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1

# ---------------------------------------------------------------- Assumptions
ws2 = wb.create_sheet("Assumptions")
ws2.column_dimensions["A"].width = 34
ws2.column_dimensions["B"].width = 14
ws2.column_dimensions["C"].width = 50
ws2["A1"] = "Model Assumptions"
ws2["A1"].font = TITLE_FONT

ws2["A3"] = "Scoring weights (must sum to 1.00)"
ws2["A3"].font = Font(name=FONT_NAME, size=11, bold=True)
weights = [
    ("Goals + Assists per 90", 0.25, "Counting production -- least reliable across levels, so weighted lowest."),
    ("Progressive actions per 90 (passes + carries)", 0.35, "Ball-progression skill -- travels well across league quality."),
    ("Defensive actions per 90 (tackles + interceptions)", 0.20, "Durable, possession-adjusted signal of work rate/defensive skill."),
    ("Youth factor (younger relative to peers scores higher)", 0.20, "A teenager holding a starting role is a stronger signal than a 26-year-old doing the same."),
]
r = 4
ws2.cell(row=r, column=1, value="Factor").font = Font(name=FONT_NAME, bold=True)
ws2.cell(row=r, column=2, value="Weight").font = Font(name=FONT_NAME, bold=True)
ws2.cell(row=r, column=3, value="Rationale").font = Font(name=FONT_NAME, bold=True)
r += 1
weight_start_row = r
for name, val, note in weights:
    ws2.cell(row=r, column=1, value=name).font = BLACK
    wc = ws2.cell(row=r, column=2, value=val)
    wc.font = BLUE
    wc.fill = YELLOW_FILL
    wc.number_format = "0%"
    nc = ws2.cell(row=r, column=3, value=note)
    nc.font = Font(name=FONT_NAME, size=9, italic=True)
    nc.alignment = Alignment(wrap_text=True)
    r += 1
weight_end_row = r - 1
ws2.cell(row=r, column=1, value="Total").font = Font(name=FONT_NAME, bold=True)
total_cell = ws2.cell(row=r, column=2, value=f"=SUM(B{weight_start_row}:B{weight_end_row})")
total_cell.font = BLACK
total_cell.number_format = "0%"

r += 2
ws2.cell(row=r, column=1, value="League tier strength multiplier").font = Font(name=FONT_NAME, size=11, bold=True)
r += 1
ws2.cell(row=r, column=1, value="Tier").font = Font(name=FONT_NAME, bold=True)
ws2.cell(row=r, column=2, value="Multiplier").font = Font(name=FONT_NAME, bold=True)
ws2.cell(row=r, column=3, value="Note").font = Font(name=FONT_NAME, bold=True)
r += 1
tier_start_row = r
tiers = [
    (2, 1.00, "2nd division -- strongest tier in scope."),
    (3, 0.85, "3rd division -- solid output here still discounted vs. tier 2."),
    (4, 0.70, "4th division -- output must be well above peers to overcome the discount."),
]
for t, mult, note in tiers:
    ws2.cell(row=r, column=1, value=t).font = BLUE
    mc = ws2.cell(row=r, column=2, value=mult)
    mc.font = BLUE
    mc.fill = YELLOW_FILL
    mc.number_format = "0.00"
    nc = ws2.cell(row=r, column=3, value=note)
    nc.font = Font(name=FONT_NAME, size=9, italic=True)
    r += 1
tier_end_row = r - 1

r += 2
ws2.cell(row=r, column=1, value="Flag thresholds (Undervalued Score, percentage points)").font = Font(name=FONT_NAME, size=11, bold=True)
r += 1
ws2.cell(row=r, column=1, value="High Priority >=").font = BLACK
hp_cell = ws2.cell(row=r, column=2, value=40)
hp_cell.font = BLUE
hp_cell.fill = YELLOW_FILL
hp_row = r
r += 1
ws2.cell(row=r, column=1, value="Watchlist >=").font = BLACK
wl_cell = ws2.cell(row=r, column=2, value=20)
wl_cell.font = BLUE
wl_cell.fill = YELLOW_FILL
wl_row = r

r += 2
ws2.cell(row=r, column=1, value="System fit thresholds (percentile, 0-1)").font = Font(name=FONT_NAME, size=11, bold=True)
r += 1
ws2.cell(row=r, column=1, value="High-percentile cutoff").font = BLACK
sf_hi_cell = ws2.cell(row=r, column=2, value=0.65)
sf_hi_cell.font = BLUE
sf_hi_cell.fill = YELLOW_FILL
sf_hi_cell.number_format = "0%"
sf_hi_row = r
r += 1
ws2.cell(row=r, column=1, value="Youth cutoff (for pressing fit)").font = BLACK
sf_youth_cell = ws2.cell(row=r, column=2, value=0.55)
sf_youth_cell.font = BLUE
sf_youth_cell.fill = YELLOW_FILL
sf_youth_cell.number_format = "0%"
sf_youth_row = r
r += 1
ws2.cell(row=r, column=1, value="Pass-driven cutoff (pass/carry ratio)").font = BLACK
sf_pass_cell = ws2.cell(row=r, column=2, value=0.62)
sf_pass_cell.font = BLUE
sf_pass_cell.fill = YELLOW_FILL
sf_pass_cell.number_format = "0%"
sf_pass_row = r
r += 1
ws2.cell(row=r, column=1, value="Low-percentile cutoff").font = BLACK
sf_lo_cell = ws2.cell(row=r, column=2, value=0.45)
sf_lo_cell.font = BLUE
sf_lo_cell.fill = YELLOW_FILL
sf_lo_cell.number_format = "0%"
sf_lo_row = r

GA_W = f"Assumptions!$B${weight_start_row}"
PROG_W = f"Assumptions!$B${weight_start_row+1}"
DEF_W = f"Assumptions!$B${weight_start_row+2}"
AGE_W = f"Assumptions!$B${weight_start_row+3}"
HP_THRESH = f"Assumptions!$B${hp_row}"
WL_THRESH = f"Assumptions!$B${wl_row}"
tier_val_range = f"Assumptions!$B${tier_start_row}:$B${tier_end_row}"
tier_key_range = f"Assumptions!$A${tier_start_row}:$A${tier_end_row}"
SF_HI = f"Assumptions!$B${sf_hi_row}"
SF_YOUTH = f"Assumptions!$B${sf_youth_row}"
SF_PASS = f"Assumptions!$B${sf_pass_row}"
SF_LO = f"Assumptions!$B${sf_lo_row}"

r += 2
ws2.cell(row=r, column=1, value="Transfer pathway countries (by player country)").font = Font(name=FONT_NAME, size=11, bold=True)
r += 1
ws2.cell(row=r, column=1, value="Country").font = Font(name=FONT_NAME, bold=True)
ws2.cell(row=r, column=2, value="Pathway Countries").font = Font(name=FONT_NAME, bold=True)
r += 1
pathway_start_row = r
for country, dests in PATHWAYS.items():
    ws2.cell(row=r, column=1, value=country).font = BLACK
    pc = ws2.cell(row=r, column=2, value=", ".join(dests))
    pc.font = BLACK
    r += 1
pathway_end_row = r - 1
pathway_key_range = f"Assumptions!$A${pathway_start_row}:$A${pathway_end_row}"
pathway_val_range = f"Assumptions!$B${pathway_start_row}:$B${pathway_end_row}"

# ---------------------------------------------------------------- Player Data
pd_ws = wb.create_sheet("Player Data")
headers = ["Player Name", "Country", "League", "Tier", "Club", "Position", "Age",
           "Minutes Played", "Goals", "Assists", "Progressive Passes", "Progressive Carries",
           "Tackles + Interceptions", "Saves (GK)", "Goals Conceded (GK)", "Pass Completion % (GK)",
           "Sweeper Actions (GK)", "Clean Sheets (GK)",
           "Market Value (EUR)", "Has Agent", "Contract Expires (Year)",
           "Club Contact Email", "Preferred Contact Route", "Age Verification Source",
           "Date Added", "Last Updated", "Verify Age / Bio"]
widths = [22, 16, 26, 7, 20, 10, 6, 14, 8, 8, 12, 12, 12, 10, 12, 14, 12, 12, 16, 11, 14, 26, 34, 34, 12, 12, 18]
for i, (h, w) in enumerate(zip(headers, widths), start=1):
    cell = pd_ws.cell(row=1, column=i, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(wrap_text=True, vertical="center")
    pd_ws.column_dimensions[get_column_letter(i)].width = w
pd_ws.row_dimensions[1].height = 32

FIELD_ORDER = ["name","country","league","tier","club","position","age","minutes","goals","assists",
               "progPasses","progCarries","tklInt","saves","goalsConceded","passCompletionPct",
               "sweeperActions","cleanSheets","marketValue","hasAgent","contractExpires",
               "clubContactEmail","contactRoute","federationRegistry","dateAdded","lastUpdated"]
VERIFY_COL = len(FIELD_ORDER) + 1

for r_idx, p in enumerate(players, start=2):
    for c_idx, field in enumerate(FIELD_ORDER, start=1):
        cell = pd_ws.cell(row=r_idx, column=c_idx, value=p[field])
        cell.font = BLUE
        cell.border = BORDER
        if field == "marketValue":
            cell.number_format = "$#,##0;($#,##0);-"
        if field == "passCompletionPct":
            cell.number_format = "0.0"
        if field in ("contactRoute", "federationRegistry"):
            cell.alignment = Alignment(wrap_text=True)
    link_cell = pd_ws.cell(row=r_idx, column=VERIFY_COL,
        value=f'=HYPERLINK("https://www.transfermarkt.com/schnellsuche/ergebnis/schnellsuche?query="&SUBSTITUTE(A{r_idx}," ","+"),"Search Transfermarkt")')
    link_cell.font = LINK
    link_cell.border = BORDER

VERIFY_LETTER = get_column_letter(VERIFY_COL)

dv_pos = DataValidation(type="list", formula1='"GK,DF,MF,FW"', allow_blank=False)
pd_ws.add_data_validation(dv_pos)
dv_pos.add(f"F2:F{last_row+20}")
dv_tier = DataValidation(type="list", formula1='"2,3,4"', allow_blank=False)
pd_ws.add_data_validation(dv_tier)
dv_tier.add(f"D2:D{last_row+20}")
dv_agent = DataValidation(type="list", formula1='"Yes,No,Unknown"', allow_blank=False)
pd_ws.add_data_validation(dv_agent)
dv_agent.add(f"T2:T{last_row+20}")

pd_ws.freeze_panes = "A2"
pd_ws.auto_filter.ref = f"A1:{VERIFY_LETTER}{last_row}"

note_row = last_row + 2
note_cell = pd_ws.cell(row=note_row, column=1,
    value=(f"Rows above are SAMPLE data ({N} fictional players, ages 17-26, spanning "
           f"{len(set(p['country'] for p in players))} countries worldwide). Club contact emails are illustrative "
           "formats, not real verified addresses -- confirm on the club's official site before use. Age "
           "Verification Source / Verify Age point to real, legitimate places to confirm identity and DOB "
           "(never a fabricated profile link, since that could accidentally point to an unrelated real person). "
           "Never contact a minor (under 18) directly -- go through the club's youth/academy office. Date Added / "
           "Last Updated reflect the weekly refresh cadence (10 new players added, all players' stats refreshed, "
           "each week -- see weekly_update.py). See the Research & Build Plan doc for real data sources."))
note_cell.font = Font(name=FONT_NAME, size=9, italic=True, color="777777")
note_cell.alignment = Alignment(wrap_text=True)
pd_ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=VERIFY_COL)
pd_ws.row_dimensions[note_row].height = 55

# ---------------------------------------------------------------- Scoring Model
sm = wb.create_sheet("Scoring Model")
sm_headers = ["Player", "Position", "Country", "League", "Tier", "Age", "Minutes",
              "GA / 90", "Prog / 90", "Def / 90", "Market Value (EUR)", "Tier Multiplier",
              "GA Percentile", "Prog Percentile", "Def Percentile", "Youth Percentile",
              "Performance Score", "Performance Percentile", "Market Visibility Percentile",
              "Undervalued Score", "Has Agent", "Contract Expires", "Flag",
              "Club Contact Email", "Preferred Contact Route", "Age Verification Source", "Verify Age / Bio",
              "Pass/Carry Ratio", "System Fit", "Date Added", "Last Updated",
              "Transfer Pathway Countries", "Move Type"]
sm_widths = [22, 9, 16, 20, 6, 6, 9, 9, 9, 9, 15, 11, 11, 11, 11, 11, 12, 12, 13, 12, 10, 12, 22, 26, 34, 34, 18, 12, 30, 12, 12, 30, 18]
for i, (h, w) in enumerate(zip(sm_headers, sm_widths), start=1):
    c = sm.cell(row=1, column=i, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = Alignment(wrap_text=True, vertical="center")
    sm.column_dimensions[get_column_letter(i)].width = w
sm.row_dimensions[1].height = 32
sm.freeze_panes = "A2"

PD = "'Player Data'!"
LAST_COL = len(sm_headers)
for i in range(2, last_row + 1):
    pd_row = i
    sm.cell(row=i, column=1, value=f"={PD}A{pd_row}").font = GREEN
    sm.cell(row=i, column=2, value=f"={PD}F{pd_row}").font = GREEN
    sm.cell(row=i, column=3, value=f"={PD}B{pd_row}").font = GREEN
    sm.cell(row=i, column=4, value=f"={PD}C{pd_row}").font = GREEN
    sm.cell(row=i, column=5, value=f"={PD}D{pd_row}").font = GREEN
    sm.cell(row=i, column=6, value=f"={PD}G{pd_row}").font = GREEN
    sm.cell(row=i, column=7, value=f"={PD}H{pd_row}").font = GREEN

    # Factor 1/2/3 are position-conditional: for GK these are Shot-Stopping
    # (saves minus a goals-conceded penalty, a simplified proxy for PSxG+/-
    # since this fictional dataset has no shot-level data), Distribution
    # (pass completion %), and Sweeper Actions/90 -- goalkeeper-appropriate
    # metrics grounded in real GK analytics (see Read Me). Outfield players
    # keep the original GA/Prog/Def formulas unchanged.
    ga = sm.cell(row=i, column=8,
        value=(f'=IF(B{i}="GK",IFERROR(({PD}N{pd_row}-1.5*{PD}O{pd_row})/{PD}H{pd_row}*90,0),'
               f'IFERROR(({PD}I{pd_row}+{PD}J{pd_row})/{PD}H{pd_row}*90,0))'))
    ga.font = GREEN; ga.number_format = "0.00"
    prog = sm.cell(row=i, column=9,
        value=(f'=IF(B{i}="GK",{PD}P{pd_row},'
               f'IFERROR(({PD}K{pd_row}+{PD}L{pd_row})/{PD}H{pd_row}*90,0))'))
    prog.font = GREEN; prog.number_format = "0.00"
    de = sm.cell(row=i, column=10,
        value=(f'=IF(B{i}="GK",IFERROR({PD}Q{pd_row}/{PD}H{pd_row}*90,0),'
               f'IFERROR({PD}M{pd_row}/{PD}H{pd_row}*90,0))'))
    de.font = GREEN; de.number_format = "0.00"

    mv = sm.cell(row=i, column=11, value=f"={PD}S{pd_row}")
    mv.font = GREEN; mv.number_format = "$#,##0;($#,##0);-"

    tm = sm.cell(row=i, column=12, value=f"=INDEX({tier_val_range},MATCH(E{i},{tier_key_range},0))")
    tm.font = GREEN; tm.number_format = "0.00"

    ga_pct = sm.cell(row=i, column=13,
        value=f"=IFERROR(SUMPRODUCT(($B$2:$B${last_row}=B{i})*($H$2:$H${last_row}<=H{i}))/SUMPRODUCT(($B$2:$B${last_row}=B{i})*1),0)")
    ga_pct.font = BLACK; ga_pct.number_format = "0%"
    prog_pct = sm.cell(row=i, column=14,
        value=f"=IFERROR(SUMPRODUCT(($B$2:$B${last_row}=B{i})*($I$2:$I${last_row}<=I{i}))/SUMPRODUCT(($B$2:$B${last_row}=B{i})*1),0)")
    prog_pct.font = BLACK; prog_pct.number_format = "0%"
    def_pct = sm.cell(row=i, column=15,
        value=f"=IFERROR(SUMPRODUCT(($B$2:$B${last_row}=B{i})*($J$2:$J${last_row}<=J{i}))/SUMPRODUCT(($B$2:$B${last_row}=B{i})*1),0)")
    def_pct.font = BLACK; def_pct.number_format = "0%"
    youth_pct = sm.cell(row=i, column=16,
        value=f"=IFERROR(SUMPRODUCT(($B$2:$B${last_row}=B{i})*((0-$F$2:$F${last_row})<=(0-F{i})))/SUMPRODUCT(($B$2:$B${last_row}=B{i})*1),0)")
    youth_pct.font = BLACK; youth_pct.number_format = "0%"

    perf = sm.cell(row=i, column=17,
        value=f"=(M{i}*{GA_W}+N{i}*{PROG_W}+O{i}*{DEF_W}+P{i}*{AGE_W})*L{i}")
    perf.font = GREEN; perf.number_format = "0.000"

    perf_pct = sm.cell(row=i, column=18,
        value=f"=IFERROR(SUMPRODUCT(($B$2:$B${last_row}=B{i})*($Q$2:$Q${last_row}<=Q{i}))/SUMPRODUCT(($B$2:$B${last_row}=B{i})*1),0)")
    perf_pct.font = BLACK; perf_pct.number_format = "0%"

    mv_pct = sm.cell(row=i, column=19,
        value=f"=IFERROR(SUMPRODUCT(($B$2:$B${last_row}=B{i})*($K$2:$K${last_row}<=K{i}))/SUMPRODUCT(($B$2:$B${last_row}=B{i})*1),0)")
    mv_pct.font = BLACK; mv_pct.number_format = "0%"

    uv = sm.cell(row=i, column=20, value=f"=(R{i}-S{i})*100")
    uv.font = BLACK; uv.number_format = "0.0"

    agent = sm.cell(row=i, column=21, value=f"={PD}T{pd_row}")
    agent.font = GREEN
    contract = sm.cell(row=i, column=22, value=f"={PD}U{pd_row}")
    contract.font = GREEN; contract.number_format = "0"

    flag = sm.cell(row=i, column=23,
        value=(f'=IF(AND(T{i}>={HP_THRESH},U{i}="No"),"High Priority - Unrepresented",'
               f'IF(T{i}>={HP_THRESH},"High Priority",'
               f'IF(AND(T{i}>={WL_THRESH},U{i}="No"),"Watchlist - Unrepresented",'
               f'IF(T{i}>={WL_THRESH},"Watchlist",""))))'))
    flag.font = GREEN

    email = sm.cell(row=i, column=24, value=f"={PD}V{pd_row}")
    email.font = GREEN
    route = sm.cell(row=i, column=25, value=f"={PD}W{pd_row}")
    route.font = GREEN; route.alignment = Alignment(wrap_text=True)
    fed = sm.cell(row=i, column=26, value=f"={PD}X{pd_row}")
    fed.font = GREEN; fed.alignment = Alignment(wrap_text=True)
    verify_link = sm.cell(row=i, column=27,
        value=f'=HYPERLINK("https://www.transfermarkt.com/schnellsuche/ergebnis/schnellsuche?query="&SUBSTITUTE(A{i}," ","+"),"Search Transfermarkt")')
    verify_link.font = LINK

    # Pass/Carry Ratio -- how much of this player's progression comes from passing vs. carrying
    pcr = sm.cell(row=i, column=28,
        value=f"=IFERROR({PD}K{pd_row}/({PD}K{pd_row}+{PD}L{pd_row}),0)")
    pcr.font = GREEN; pcr.number_format = "0%"

    # System Fit -- rules-based tactical classification from the percentiles already computed above
    sysfit = sm.cell(row=i, column=29,
        value=(f'=IF(B{i}="GK",'
               f'IF(AND(O{i}>={SF_HI},N{i}>={SF_HI}),"Sweeper-Keeper / Build-from-the-Back",'
               f'IF(O{i}>={SF_HI},"Sweeper-Keeper / High Line",'
               f'IF(M{i}>={SF_HI},"Shot-Stopper / Traditional",'
               f'"Flexible / Multi-System"))),'
               f'IF(AND(O{i}>={SF_HI},P{i}>={SF_YOUTH}),"High Press / Gegenpressing",'
               f'IF(AND(N{i}>={SF_HI},AB{i}>={SF_PASS}),"Possession / Build-from-the-Back",'
               f'IF(N{i}>={SF_HI},"Counter-Attack / Transition",'
               f'IF(AND(O{i}>={SF_HI},N{i}<{SF_LO}),"Low Block / Park the Bus",'
               f'IF(AND(B{i}="FW",M{i}>={SF_HI},N{i}<{SF_LO}),"Direct / Target Approach",'
               f'"Flexible / Multi-System"))))))'))
    sysfit.font = GREEN

    added = sm.cell(row=i, column=30, value=f"={PD}Y{pd_row}")
    added.font = GREEN
    updated = sm.cell(row=i, column=31, value=f"={PD}Z{pd_row}")
    updated.font = GREEN

    pathway = sm.cell(row=i, column=32,
        value=f"=IFERROR(INDEX({pathway_val_range},MATCH(C{i},{pathway_key_range},0)),\"--\")")
    pathway.font = GREEN; pathway.alignment = Alignment(wrap_text=True)

    move_type = sm.cell(row=i, column=33,
        value=(f'=IF(R{i}>=0.7,"Step-Up Candidate",IF(R{i}>=0.4,"Lateral Move","Stay & Develop"))'))
    move_type.font = BLACK

    for col in range(1, LAST_COL + 1):
        sm.cell(row=i, column=col).border = BORDER

sm.auto_filter.ref = f"A1:{get_column_letter(LAST_COL)}{last_row}"

color_rule = ColorScaleRule(
    start_type="min", start_color="F8696B",
    mid_type="percentile", mid_value=50, mid_color="FFEB84",
    end_type="max", end_color="63BE7B",
)
sm.conditional_formatting.add(f"T2:T{last_row}", color_rule)

hp_unrep_rule = CellIsRule(operator="equal", formula=['"High Priority - Unrepresented"'], fill=PatternFill("solid", fgColor="00B050"), font=Font(name=FONT_NAME, color="FFFFFF", bold=True))
hp_rule = CellIsRule(operator="equal", formula=['"High Priority"'], fill=PatternFill("solid", fgColor="C6EFCE"), font=Font(name=FONT_NAME, color="006100", bold=True))
wl_unrep_rule = CellIsRule(operator="equal", formula=['"Watchlist - Unrepresented"'], fill=PatternFill("solid", fgColor="FFD966"), font=Font(name=FONT_NAME, color="7F6000", bold=True))
wl_rule = CellIsRule(operator="equal", formula=['"Watchlist"'], fill=PatternFill("solid", fgColor="FFEB9C"), font=Font(name=FONT_NAME, color="9C6500"))
sm.conditional_formatting.add(f"W2:W{last_row}", hp_unrep_rule)
sm.conditional_formatting.add(f"W2:W{last_row}", hp_rule)
sm.conditional_formatting.add(f"W2:W{last_row}", wl_unrep_rule)
agent_rule = CellIsRule(operator="equal", formula=['"No"'], fill=PatternFill("solid", fgColor="E2EFDA"))
sm.conditional_formatting.add(f"U2:U{last_row}", agent_rule)

wb.save(OUT_FILE)
print(f"saved {OUT_FILE} with {N} players, last_row={last_row}")
