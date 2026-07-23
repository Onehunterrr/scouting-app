"""
Regenerates test_app.js from the CURRENT state of players_current.json and the
freshly recalculated Scouting_Model.xlsx. This makes the test suite safe to run
after every weekly update: it never hardcodes a player count, country count, or
score/flag value that would go stale as the database grows and refreshes.
"""
import json, sys
from openpyxl import load_workbook

DATA_FILE = sys.argv[1] if len(sys.argv) > 1 else "players_current.json"
XLSX_FILE = sys.argv[2] if len(sys.argv) > 2 else "Scouting_Model.xlsx"
OUT_FILE = sys.argv[3] if len(sys.argv) > 3 else "test_app.js"

players = json.load(open(DATA_FILE))
total_players = len(players)
countries = sorted(set(p["country"] for p in players))

wb = load_workbook(XLSX_FILE, data_only=True)
sm = wb["Scoring Model"]
computed = {}
for r in range(2, sm.max_row + 1):
    name = sm.cell(row=r, column=1).value
    if not name:
        continue
    computed[name] = {
        "undervaluedScore": sm.cell(row=r, column=20).value,
        "flag": sm.cell(row=r, column=23).value or "",
        "systemFit": sm.cell(row=r, column=29).value,
    }

by_name = {p["name"]: p for p in players}

# ---- pick dynamic test subjects ----
minor = next((p for p in players if p["age"] < 18), None)
agented = next((p for p in players if p["hasAgent"] == "Yes"), None)
unrep = next((p for p in players if p["hasAgent"] == "No" and computed.get(p["name"], {}).get("flag", "").endswith("Unrepresented")), None)

# a player with a search token that is unique across the WHOLE search haystack
# (the app's search now matches name OR club OR country, so the token must not
# appear in any other player's name, club, or country either)
search_target = None
haystacks = [(p["name"], (p["name"] + " " + p["club"] + " " + p["country"]).lower()) for p in players]
last_tokens = {}
for p in players:
    tok = p["name"].split()[-1].lower()
    last_tokens.setdefault(tok, []).append(p["name"])
for tok, names in last_tokens.items():
    if len(names) == 1 and len(tok) >= 4:
        hits = sum(1 for _, hay in haystacks if tok in hay)
        if hits == 1:
            search_target = names[0]
            search_token = tok
            break

# 4 arbitrary cross-check players spread across the flag spectrum, by Undervalued Score
scored_sorted = sorted(players, key=lambda p: computed.get(p["name"], {}).get("undervaluedScore", 0) or 0, reverse=True)
cross_check_names = []
for p in [scored_sorted[0], scored_sorted[len(scored_sorted)//3], scored_sorted[2*len(scored_sorted)//3], scored_sorted[-1]]:
    if p["name"] not in cross_check_names:
        cross_check_names.append(p["name"])

cross_checks = {n: computed[n]["undervaluedScore"] for n in cross_check_names}
flag_checks = {n: computed[n]["flag"] for n in cross_check_names if computed[n]["flag"]}

# system fit cross-check: any player whose System Fit isn't the generic fallback
system_fit_subject = next((p["name"] for p in players
                            if computed.get(p["name"], {}).get("systemFit") not in (None, "Flexible / Multi-System")), None)
if system_fit_subject is None:
    system_fit_subject = players[0]["name"]
system_fit_expected = computed[system_fit_subject]["systemFit"]

# tier that actually has players, for the tier filter test
tiers_present = sorted(set(p["tier"] for p in players))
tier_test = tiers_present[-1] if tiers_present else 4

latest_added = max(p["dateAdded"] for p in players)
new_batch_count = sum(1 for p in players if p["dateAdded"] == latest_added)

country_option_count = len(countries) + 1  # + "All countries"

minor_block = "" if not minor else f'''
  // 18. Minor-safety warning
  const minorRow = rows.find(r => r.querySelector("td.name").textContent.includes("{minor['name']}"));
  assert(minorRow !== undefined, "minor test subject exists in table");
  minorRow.dispatchEvent(new w.Event("click", {{ bubbles: true }}));
  const minorContactHtml = document.getElementById("m-contact").innerHTML;
  assert(minorContactHtml.includes("minor-warning"), "minor warning block rendered for a player under 18");
  assert(minorContactHtml.includes("youth/academy office"), "minor route points to club youth office, not direct contact");
  document.getElementById("modal-close").dispatchEvent(new w.Event("click", {{ bubbles: true }}));
'''

agented_block = "" if not agented else f'''
  // 19. Agent-routing
  const agentedRow = rows.find(r => r.querySelector("td.name").textContent.includes("{agented['name']}"));
  assert(agentedRow !== undefined, "agented test subject exists in table");
  agentedRow.dispatchEvent(new w.Event("click", {{ bubbles: true }}));
  const agentedContactHtml = document.getElementById("m-contact").innerHTML;
  assert(agentedContactHtml.includes("player's agent"), "agented player routes through their agent");
  document.getElementById("modal-close").dispatchEvent(new w.Event("click", {{ bubbles: true }}));
'''

test_js = f'''const fs = require("fs");
const path = require("path");
const {{ JSDOM }} = require("jsdom");

const html = fs.readFileSync(path.join(__dirname, "Scouting_App_Prototype.html"), "utf8");

function newDom() {{
  const errors = [];
  const dom = new JSDOM(html, {{
    runScripts: "dangerously",
    resources: "usable",
    pretendToBeVisual: true,
  }});
  dom.window.onerror = (msg) => errors.push(msg);
  return {{ dom, errors }};
}}

function assert(cond, msg) {{
  if (!cond) throw new Error("ASSERTION FAILED: " + msg);
  console.log("  OK: " + msg);
}}

function rowCount(doc) {{
  return doc.querySelectorAll("#table-body tr").length;
}}

function runPass(passNum) {{
  console.log(`\\n=== PASS ${{passNum}} ===`);
  const {{ dom, errors }} = newDom();
  const {{ document }} = dom.window;
  const w = dom.window;

  const totalPlayers = {total_players};

  // 0. Pagination default: offline mode starts at 50/page, page 1
  const expectedFirstPage = Math.min(50, totalPlayers);
  assert(rowCount(document) === expectedFirstPage, `initial render shows first page of ${{expectedFirstPage}} rows (50/page default)`);
  assert(document.getElementById("stat-count").textContent === String(totalPlayers), "stat-count shows the FULL filtered total, not just the page");
  assert(document.getElementById("page-info").textContent.startsWith("Page 1 of"), "pager shows 'Page 1 of ...' on load");

  // switch the page size to "All" (offline-only option) so every subsequent
  // assertion sees the complete result set, exactly like the pre-pagination app
  const pageSizeSel = document.getElementById("page-size");
  pageSizeSel.value = "all";
  pageSizeSel.dispatchEvent(new w.Event("change", {{ bubbles: true }}));

  // 1. Initial render
  assert(rowCount(document) === totalPlayers, `initial render shows all ${{totalPlayers}} players`);
  assert(document.getElementById("stat-count").textContent === String(totalPlayers), "stat-count matches row count");
  const countryOptions = document.querySelectorAll("#f-country option").length;
  assert(countryOptions === {country_option_count}, `country filter populated ({len(countries)} countries + 'All' = {country_option_count} options), got ${{countryOptions}}`);

  // 2. Default sort is Undervalued Score desc -> first row numeric
  const firstRowScoreText = document.querySelector("#table-body tr td.uv-score").textContent;
  assert(!isNaN(parseFloat(firstRowScoreText)), "first row undervalued score is numeric: " + firstRowScoreText);

  // 3. Name sort toggle
  const nameHeader = [...document.querySelectorAll("thead th")].find(th => th.dataset.key === "name");
  nameHeader.dispatchEvent(new w.Event("click", {{ bubbles: true }}));
  const firstNameDesc = document.querySelector("#table-body tr td.name").textContent;
  nameHeader.dispatchEvent(new w.Event("click", {{ bubbles: true }}));
  const firstNameAsc = document.querySelector("#table-body tr td.name").textContent;
  assert(firstNameAsc !== firstNameDesc, `sort toggle changes order (desc first="${{firstNameDesc}}", asc first="${{firstNameAsc}}")`);

  const uvHeader = [...document.querySelectorAll("thead th")].find(th => th.dataset.key === "undervaluedScore");
  uvHeader.dispatchEvent(new w.Event("click", {{ bubbles: true }}));

  // 4. Position filter
  const posSelect = document.getElementById("f-position");
  posSelect.value = "FW";
  posSelect.dispatchEvent(new w.Event("input", {{ bubbles: true }}));
  const fwRows = rowCount(document);
  assert(fwRows > 0 && fwRows < totalPlayers, `position filter (FW) narrows results: ${{fwRows}} rows`);
  const allFW = [...document.querySelectorAll("#table-body tr")].every(tr => tr.children[3].textContent === "FW");
  assert(allFW, "all visible rows are FW after position filter");

  // 5. Country filter stacked on position
  const countrySelect = document.getElementById("f-country");
  const someCountry = [...document.querySelectorAll("#f-country option")][1].value;
  countrySelect.value = someCountry;
  countrySelect.dispatchEvent(new w.Event("input", {{ bubbles: true }}));
  const stackedRows = rowCount(document);
  assert(stackedRows <= fwRows, `country filter further narrows results: ${{stackedRows}} rows`);

  // 6. Reset filters
  document.getElementById("reset-filters").dispatchEvent(new w.Event("click", {{ bubbles: true }}));
  assert(rowCount(document) === totalPlayers, "reset filters restores all players");

  // 7. Search filter (dynamically chosen unique name from current dataset)
  const search = document.getElementById("f-search");
  search.value = "{search_token if search_target else 'zzz_no_match'}";
  search.dispatchEvent(new w.Event("input", {{ bubbles: true }}));
  assert(rowCount(document) === 1, `search '{search_token if search_target else 'zzz_no_match'}' returns exactly 1 row, got ${{rowCount(document)}}`);
  assert(document.querySelector("#table-body tr td.name").textContent.includes("{search_target or ''}"), "search result is correct player");
  search.value = "";
  search.dispatchEvent(new w.Event("input", {{ bubbles: true }}));

  // 8. Age slider filter
  const ageSlider = document.getElementById("f-age");
  ageSlider.value = "18";
  ageSlider.dispatchEvent(new w.Event("input", {{ bubbles: true }}));
  const allUnder18 = [...document.querySelectorAll("#table-body tr")].every(tr => parseInt(tr.children[6].textContent, 10) <= 18);
  assert(allUnder18, "age filter (<=18) respected");
  ageSlider.value = "26";
  ageSlider.dispatchEvent(new w.Event("input", {{ bubbles: true }}));

  // 9. Agent checkbox filter
  const noCb = [...document.querySelectorAll(".agent-cb")].find(cb => cb.value === "No");
  noCb.checked = false;
  noCb.dispatchEvent(new w.Event("change", {{ bubbles: true }}));
  const noAgentHidden = [...document.querySelectorAll("#table-body tr")].every(tr => {{
    const tag = tr.querySelector(".agent-tag");
    return tag.textContent.trim() !== "No";
  }});
  assert(noAgentHidden, "unchecking 'No agent' hides all unrepresented players");
  noCb.checked = true;
  noCb.dispatchEvent(new w.Event("change", {{ bubbles: true }}));
  assert(rowCount(document) === totalPlayers, "re-checking restores all players");

  // 10. Weight sliders change scores
  const gaSlider = document.getElementById("w-ga");
  gaSlider.value = "80";
  gaSlider.dispatchEvent(new w.Event("input", {{ bubbles: true }}));
  const progSlider = document.getElementById("w-prog");
  progSlider.value = "5";
  progSlider.dispatchEvent(new w.Event("input", {{ bubbles: true }}));
  assert(document.getElementById("w-ga-val").textContent === "80%", "weight label updates to 80%");

  // 11. Reset weights
  document.getElementById("reset-weights").dispatchEvent(new w.Event("click", {{ bubbles: true }}));
  assert(document.getElementById("w-ga-val").textContent === "25%", "reset weights restores GA to 25%");

  // 12. Open modal
  const firstRow = document.querySelector("#table-body tr");
  const clickedName = firstRow.querySelector("td.name").textContent.replace("NEW", "").trim();
  firstRow.dispatchEvent(new w.Event("click", {{ bubbles: true }}));
  assert(document.getElementById("modal-overlay").classList.contains("open"), "modal opens on row click");
  assert(document.getElementById("m-name").textContent === clickedName, "modal shows correct player name: " + clickedName);
  const barBlocks = document.querySelectorAll("#m-bars .bar-block").length;
  assert(barBlocks === 6, `modal shows 6 stat bars, got ${{barBlocks}}`);
  const explainText = document.getElementById("m-explain").textContent;
  assert(explainText.includes("Undervalued Score"), "modal explanation mentions Undervalued Score");

  // 12b. Date Added / Last Updated pills
  const pillsHtml = document.getElementById("m-pills").innerHTML;
  assert(pillsHtml.includes("Added"), "Date Added pill present");
  assert(pillsHtml.includes("Updated"), "Last Updated pill present");

  // 12c. Tactical System Fit section (separate from Contact & Verification)
  const systemHtml = document.getElementById("m-system").innerHTML;
  assert(systemHtml.includes("Tactical System Fit"), "system fit section header present");
  assert(document.querySelector("#m-system .system-badge") !== null, "system fit badge rendered");
  assert(document.querySelector("#m-system .system-note").textContent.length > 20, "system fit rationale text rendered");

  // 12d. Contact & Verification section
  const contactHtml = document.getElementById("m-contact").innerHTML;
  assert(contactHtml.includes("Contact &amp; Verification"), "contact section header present");
  const mailLink = document.querySelector("#m-contact a[href^='mailto:']");
  assert(mailLink !== null, "club contact email mailto link present");
  const tmLink = document.querySelector("#m-contact a[href*='transfermarkt.com']");
  assert(tmLink !== null, "Transfermarkt verification link present");
  assert(tmLink.getAttribute("href").includes("schnellsuche"), "Transfermarkt link is a live search, not a fabricated profile URL");
  assert(contactHtml.includes("Preferred route"), "preferred contact route shown");
  assert(contactHtml.includes("Also check"), "federation registry cross-check shown");

  // 13. Close modal via close button
  document.getElementById("modal-close").dispatchEvent(new w.Event("click", {{ bubbles: true }}));
  assert(!document.getElementById("modal-overlay").classList.contains("open"), "modal closes via close button");

  // 14. Open modal again, close via overlay click
  firstRow.dispatchEvent(new w.Event("click", {{ bubbles: true }}));
  const overlay = document.getElementById("modal-overlay");
  const overlayClickEvent = new w.MouseEvent("click", {{ bubbles: true }});
  Object.defineProperty(overlayClickEvent, "target", {{ value: overlay }});
  overlay.dispatchEvent(overlayClickEvent);
  assert(!overlay.classList.contains("open"), "modal closes via overlay click");

  // 15. Tier filter
  const tierSelect = document.getElementById("f-tier");
  tierSelect.value = "{tier_test}";
  tierSelect.dispatchEvent(new w.Event("input", {{ bubbles: true }}));
  const tierRows = rowCount(document);
  const allTierMatch = [...document.querySelectorAll("#table-body tr")].every(tr => tr.children[5].textContent === "{tier_test}");
  assert(allTierMatch && tierRows > 0, `tier filter ({tier_test}) works, ${{tierRows}} rows all tier {tier_test}`);
  document.getElementById("reset-filters").dispatchEvent(new w.Event("click", {{ bubbles: true }}));

  // 16. Cross-check computed Undervalued Score against known spreadsheet values (from recalculated .xlsx)
  const expected = {json.dumps(cross_checks)};
  const rows = [...document.querySelectorAll("#table-body tr")];
  for (const [name, exp] of Object.entries(expected)) {{
    const tr = rows.find(r => r.querySelector("td.name").textContent.includes(name));
    assert(tr !== undefined, `player "${{name}}" exists in rendered table`);
    const got = parseFloat(tr.querySelector("td.uv-score").textContent);
    assert(Math.abs(got - exp) < 0.5, `${{name}}: undervalued score ${{got}} matches spreadsheet ${{exp}}`);
  }}

  // 17. Flag cross-check
  const flagExpected = {json.dumps(flag_checks)};
  for (const [name, exp] of Object.entries(flagExpected)) {{
    const tr = rows.find(r => r.querySelector("td.name").textContent.includes(name));
    const badge = tr.querySelector(".badge");
    assert(badge.textContent === exp, `${{name}}: flag "${{badge.textContent}}" matches expected "${{exp}}"`);
  }}

  // 17b. System Fit cross-check against recalculated spreadsheet
  const sysRow = rows.find(r => r.querySelector("td.name").textContent.includes("{system_fit_subject}"));
  assert(sysRow !== undefined, "system fit test subject exists in table");
  sysRow.dispatchEvent(new w.Event("click", {{ bubbles: true }}));
  const sysBadgeText = document.querySelector("#m-system .system-badge").textContent;
  assert(sysBadgeText === "{system_fit_expected}", `system fit "${{sysBadgeText}}" matches spreadsheet "{system_fit_expected}"`);
  document.getElementById("modal-close").dispatchEvent(new w.Event("click", {{ bubbles: true }}));

  {minor_block}
  {agented_block}

  // 20. Country filter option count
  const allCountryOptions = [...document.querySelectorAll("#f-country option")].map(o => o.value).filter(Boolean);
  assert(allCountryOptions.length === {len(countries)}, `{len(countries)} countries present in filter, got ${{allCountryOptions.length}}`);

  // 21. Weekly-growth "NEW" badge: players from the latest batch are tagged
  const newBadges = document.querySelectorAll(".new-badge").length;
  assert(newBadges === {new_batch_count}, `${{newBadges}} NEW badges rendered, expected {new_batch_count} (players added ${{"{latest_added}"}})`);
  const statNew = document.getElementById("stat-new").textContent;
  assert(statNew === "{new_batch_count}", `stat-new chip shows {new_batch_count}, got ${{statNew}}`);

  // 22. Realistic Transfer Targets section
  const anyRow = document.querySelector("#table-body tr");
  anyRow.dispatchEvent(new w.Event("click", {{ bubbles: true }}));
  const transferHtml = document.getElementById("m-transfer").innerHTML;
  assert(transferHtml.includes("Realistic Transfer Targets"), "transfer targets section header present");
  const transferItems = document.querySelectorAll("#m-transfer .transfer-item").length;
  assert(transferItems === 5, `transfer targets shows exactly 5 candidates, got ${{transferItems}}`);
  const transferRanks = [...document.querySelectorAll("#m-transfer .transfer-rank")].map(el => el.textContent);
  assert(JSON.stringify(transferRanks) === JSON.stringify(["1","2","3","4","5"]), `transfer targets ranked 1-5, got ${{transferRanks}}`);
  const fitPercents = [...document.querySelectorAll("#m-transfer .transfer-fit b")].map(el => parseInt(el.textContent, 10));
  const isSortedDesc = fitPercents.every((v, i) => i === 0 || fitPercents[i-1] >= v);
  assert(isSortedDesc, `transfer targets sorted by fit % descending: ${{fitPercents}}`);
  const moveTypeBadges = document.querySelectorAll("#m-transfer .move-type-badge").length;
  assert(moveTypeBadges === 5, "every transfer target has a move-type badge");
  document.getElementById("modal-close").dispatchEvent(new w.Event("click", {{ bubbles: true }}));

  // 22b. Transfer targets are deterministic (stable across modal re-opens for the same player)
  anyRow.dispatchEvent(new w.Event("click", {{ bubbles: true }}));
  const clubsFirstOpen = [...document.querySelectorAll("#m-transfer .transfer-club")].map(el => el.textContent);
  document.getElementById("modal-close").dispatchEvent(new w.Event("click", {{ bubbles: true }}));
  anyRow.dispatchEvent(new w.Event("click", {{ bubbles: true }}));
  const clubsSecondOpen = [...document.querySelectorAll("#m-transfer .transfer-club")].map(el => el.textContent);
  assert(JSON.stringify(clubsFirstOpen) === JSON.stringify(clubsSecondOpen), "transfer targets are stable across repeated modal opens");
  document.getElementById("modal-close").dispatchEvent(new w.Event("click", {{ bubbles: true }}));

  // JS runtime errors check
  assert(errors.length === 0, `no uncaught JS errors during pass (${{errors.length}} found: ${{errors.join("; ")}})`);

  dom.window.close();
  console.log(`=== PASS ${{passNum}} COMPLETE -- all checks passed ===`);
}}

/* Run all 3 passes by default; `node test_app.js N` runs pass N alone so the
   three passes can be executed as three separate processes on slow machines
   (identical coverage -- each pass is a fully fresh DOM either way). */
const onlyPass = parseInt(process.argv[2] || "0", 10);
if (onlyPass >= 1 && onlyPass <= 3) {{
  runPass(onlyPass);
  console.log(`\\nPASS ${{onlyPass}} PASSED.`);
}} else {{
  for (let i = 1; i <= 3; i++) {{
    runPass(i);
  }}
  console.log("\\nALL 3 PASSES PASSED.");
}}
'''

open(OUT_FILE, "w").write(test_js)
print(f"wrote {OUT_FILE}: totalPlayers={total_players}, countries={len(countries)}, "
      f"minor={minor['name'] if minor else None}, agented={agented['name'] if agented else None}, "
      f"search={search_target}, cross_checks={cross_check_names}, system_fit_subject={system_fit_subject}, "
      f"tier_test={tier_test}, new_batch={new_batch_count}")
